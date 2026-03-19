"""
Music For All - Recap PDF → Excel Extractor  (Era-Segmented Edition)
=====================================================================
Reads all recap PDFs and produces a single Excel file with one row per
band performance.

Era breakdown
─────────────
  ERA_CLASSIC  1976–1984   Older BOA format: 13-col scores, Class + Place in Class, no Rating
  ERA_MID      1985–1999   Similar but may add Rating column
  ERA_2000s    2000–2012   Adds separate state column, Rating I/II/III, Overall place
  ERA_MODERN   2013–2019   CompetitionSuite PDFs, Rating, Overall place, size class (AA/AAA)
  ERA_2020s    2020–2025   Same as modern but slightly different column ordering on some events

Usage:
    python parse_recaps_to_excel.py
    python parse_recaps_to_excel.py --input musicforall_recaps --output recaps.xlsx
"""

import re
import sys
import logging
import argparse
from pathlib import Path
from datetime import datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Optional OCR support
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

import json
import urllib.request
import urllib.parse

# ── ArcGIS geocoding ──────────────────────────────────────────────────────────
ARCGIS_GEOCODE_URL = (
    "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates"
)
GEOCACHE_FILE = Path("venue_geocache.json")

def load_geocache() -> dict:
    if GEOCACHE_FILE.exists():
        try:
            return json.loads(GEOCACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_geocache(cache: dict):
    GEOCACHE_FILE.write_text(json.dumps(cache, indent=2), encoding="utf-8")

def geocode_venue(venue_name: str, competition: str, cache: dict) -> dict:
    city_hint = ""
    cm = re.search(r"at\s+(.+?)(?:,\s*[A-Z]{2})?$", competition, re.IGNORECASE)
    if cm:
        city_hint = cm.group(1).strip()
    search_text = f"{venue_name}, {city_hint}".strip(", ")
    cache_key   = search_text.lower()
    if cache_key in cache:
        return cache[cache_key]
    result = {"street": "", "city": "", "state": "", "zip": ""}
    try:
        params = urllib.parse.urlencode({
            "SingleLine": search_text, "f": "json", "maxLocations": 1,
            "outFields": "StAddr,City,RegionAbbr,Postal",
        })
        url = f"{ARCGIS_GEOCODE_URL}?{params}"
        req = urllib.request.Request(url, headers={"User-Agent": "MFARecapParser/2.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        candidates = data.get("candidates", [])
        if candidates:
            attrs = candidates[0].get("attributes", {})
            result = {
                "street": attrs.get("StAddr", ""),
                "city":   attrs.get("City", ""),
                "state":  attrs.get("RegionAbbr", ""),
                "zip":    attrs.get("Postal", ""),
            }
            log.info("    Geocoded '%s' → %s, %s %s", search_text,
                     result["city"], result["state"], result["zip"])
        else:
            log.warning("    No geocode results for '%s'", search_text)
    except Exception as exc:
        log.warning("    Geocode failed for '%s': %s", search_text, exc)
    cache[cache_key] = result
    return result

# ── Logging ───────────────────────────────────────────────────────────────────
LOG_FILE = "parse_log.txt"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ── Era definitions ───────────────────────────────────────────────────────────
ERA_CLASSIC = "Classic (1976-1984)"
ERA_MID     = "Mid (1985-1999)"
ERA_2000s   = "2000s (2000-2012)"
ERA_MODERN  = "Modern (2013-2019)"
ERA_2020s   = "Recent (2020-2025)"

def get_era(year: int) -> str:
    if year <= 1984: return ERA_CLASSIC
    if year <= 1999: return ERA_MID
    if year <= 2012: return ERA_2000s
    if year <= 2019: return ERA_MODERN
    return ERA_2020s

# ── Output columns ────────────────────────────────────────────────────────────
# GE column semantics differ by era:
#   1976–2012: Mus 1 GE + Mus 2 GE → Total Mus GE;  March/Vis GE = separate Vis GE judge;
#              GE Total = Total Mus GE + March/Vis GE  (sum of both judges)
#   2013–2025: BOA dropped the Vis GE sub-judge. "GE Total" now = single standalone GE
#              judge score (~14–18). "GE Vis (Standalone)" carries that value so it is
#              distinguishable from the pre-2013 summed GE Total.
#              March/Vis GE = blank (column does not exist in source PDFs).
COLUMNS = [
    "Date", "Year", "Era",
    "Competition", "Championship Type", "Performance",
    "Venue Name", "Street Address", "City", "State", "Zip",
    "School", "School State",
    "Inv Mus", "Ens Mus", "Avg Mus",
    "Inv Vis", "Ens Vis", "Avg Vis",
    "Mus 1 GE", "Mus 2 GE", "Total Mus GE", "March/Vis GE", "GE Total",
    "GE Vis (Standalone)",   # 2013+: single GE judge score; blank for pre-2013
    "Subtotal", "Penalty", "Total",
    "Place in Class", "Class", "Place Overall", "Place in Panel", "Rating",
    "Source File",
]

# ── Regex helpers ─────────────────────────────────────────────────────────────
RE_TITLE       = re.compile(r"^\d{4}\s+.+", re.IGNORECASE)
RE_DATE        = re.compile(
    r"(?:"
    r"(?:January|February|March|April|May|June|July|August|September|October|November|December)"
    r"[\s•·]+\d{1,2}(?:[,\-–]\s*\d{1,2})?,\s*\d{4}"
    r"|"
    r"\d{1,2}/\d{1,2}/\d{4}"
    r"|"
    r"\d{1,2}/\d{1,2}/(?:6\d|7\d|8\d|9\d)"
    r")",
    re.IGNORECASE,
)
RE_DATE_MMDD   = re.compile(r"(?<![\d/])\d{1,2}/\d{1,2}(?!/\d)")
RE_PERF        = re.compile(
    r"(prelim(?:inary|s)?(?:\s+\d)?|final(?:s)?|semi.?final(?:s)?|quarter.?final(?:s)?)",
    re.IGNORECASE,
)
RE_CHAMP       = re.compile(
    r"(grand\s+national|super\s+regional|regional|invitational|national|championship)",
    re.IGNORECASE,
)
RE_STATE_CELL  = re.compile(r"^[A-Z]{2}$")
RE_STATE_TRAIL = re.compile(r",?\s+([A-Z]{2})\s*$")
RE_STATE_DASH  = re.compile(r"\s+-\s+([A-Z]{2})\s*$")
RE_NUM         = re.compile(r"^\d{1,3}(?:\.\d{1,3})?$")
RE_CLASS       = re.compile(r"^(Open|A{1,3}|W|[1-6]A|AAAA?)$", re.IGNORECASE)
RE_PLACE       = re.compile(r"^\d{1,3}$")
RE_ROMAN       = re.compile(r"^(I{1,3}|IV|VI{0,3})$")
RE_RATING_DIGIT= re.compile(r"^[1-5]$")
RE_PANEL_NUM   = re.compile(r"prelim(?:s|inary)?\s+(\d)", re.IGNORECASE)
_CID_RE        = re.compile(r'\(cid:(\d+)\)')
_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f\xad\ufffe\uffff]")


def _decode_cid(text: str) -> str:
    def sub(m):
        c = chr(int(m.group(1)) + 29)
        return c if c.isprintable() else " "
    return _CID_RE.sub(sub, text)


def num_or_none(val: str):
    v = val.strip()
    if RE_NUM.match(v) and not RE_STATE_CELL.match(v):
        return float(v)
    return None


def sanitize(v):
    if isinstance(v, str):
        return _ILLEGAL_CHARS.sub("", v).strip()
    return v


def normalise_performance(raw: str) -> tuple:
    m = RE_PERF.search(raw)
    if not m:
        return "", ""
    base  = m.group(1).strip()
    base  = base[0].upper() + base[1:].lower()
    pm    = RE_PANEL_NUM.search(raw)
    panel = pm.group(1) if pm else ""
    label = f"{base} {panel}".strip() if panel else base
    return label, panel


def _strip_leading_place(s: str) -> str:
    return re.sub(r"^\d{1,3}\s+", "", s).strip()


_RE_GARBLED_SCORE = re.compile(r'\s+[A-Z]\d+[A-Z]\d+\.\d+\s*$')  # Fix 10
_RE_TRAILING_PEN  = re.compile(r'\s+(-\d+(?:\.\d+)?)\s*$')          # Fix 4

def parse_school_and_state(tokens: list) -> tuple:
    """
    Returns (school_name, state_abbr, extracted_penalty_or_None).

    FIX 4: 2023+ PDFs render "School - ST  -0.30" (penalty) in one cell.
            We extract and return the penalty so the caller can store it.
    FIX 10: Strip garbled OCR score fragments (e.g. "V1A4.700") from long school names.
    """
    if not tokens:
        return "", "", None

    joined = " ".join(tokens)

    # Fix 10: remove garbled score fragments appended to long school names
    joined = _RE_GARBLED_SCORE.sub("", joined).strip()

    # Fix 4: extract and remove trailing penalty like " -0.30"
    extracted_penalty = None
    pm = _RE_TRAILING_PEN.search(joined)
    if pm:
        try:
            extracted_penalty = abs(float(pm.group(1)))
        except ValueError:
            pass
        joined = joined[:pm.start()].strip()

    tokens = joined.split()
    if not tokens:
        return "", "", extracted_penalty

    if len(tokens) >= 2 and RE_STATE_CELL.match(tokens[-1]):
        state  = tokens[-1]
        school = " ".join(tokens[:-1]).strip().rstrip(",- ").strip()
        return _strip_leading_place(school), state, extracted_penalty

    m = RE_STATE_DASH.search(joined)
    if m:
        return _strip_leading_place(joined[:m.start()].strip()), m.group(1), extracted_penalty

    m = RE_STATE_TRAIL.search(joined)
    if m:
        return _strip_leading_place(joined[:m.start()].strip().rstrip(",").strip()), m.group(1), extracted_penalty

    return _strip_leading_place(joined), "", extracted_penalty


# ── PDF text + table extraction ───────────────────────────────────────────────

def extract_text(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            parts = []
            for page in pdf.pages:
                t = page.extract_text(x_tolerance=3, y_tolerance=3)
                if t:
                    parts.append(t)
            text = "\n".join(parts)
            if text.strip():
                if "(cid:" in text:
                    text = _decode_cid(text)
                return text
    except Exception as exc:
        log.warning("pdfplumber text extract failed %s: %s", pdf_path, exc)
    if OCR_AVAILABLE:
        log.info("    Trying OCR: %s", pdf_path.name)
        try:
            images = convert_from_path(str(pdf_path), dpi=300)
            return "\n".join(pytesseract.image_to_string(img) for img in images)
        except Exception as exc:
            log.warning("    OCR failed %s: %s", pdf_path.name, exc)
    else:
        log.warning("    Scanned PDF (no OCR available): %s", pdf_path.name)
    return ""


def extract_table_rows(pdf_path: Path) -> list:
    rows = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables({
                    "vertical_strategy":   "lines_strict",
                    "horizontal_strategy": "lines_strict",
                    "snap_tolerance": 5, "join_tolerance": 3,
                    "edge_min_length": 10,
                    "min_words_vertical": 1, "min_words_horizontal": 1,
                })
                if not tables:
                    tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        cleaned = [_decode_cid(str(c or "")).replace("\n", " ").strip() for c in row]
                        if any(cleaned):
                            rows.append(cleaned)
                if not rows:
                    text = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
                    for line in text.splitlines():
                        tokens = line.split()
                        if len(tokens) >= 8:
                            rows.append(tokens)
    except Exception as exc:
        log.warning("Table extract failed %s: %s", pdf_path, exc)
    return rows


# ── Header parsing ────────────────────────────────────────────────────────────

def parse_header(lines: list, pdf_path: Path) -> dict:
    meta = {
        "date": "", "year": 0, "era": "",
        "competition": "", "championship_type": "",
        "performance": "", "panel": "",
        "raw_title": "", "venue": "",
    }

    header_text = " ".join(lines[:20])

    # ── Infer year ────────────────────────────────────────────────────────────
    inferred_year = ""
    for line in lines[:5]:
        m = re.match(r"^(\d{4})\s+", line.strip())
        if m:
            inferred_year = m.group(1); break
    if not inferred_year:
        m = re.search(r"(?:^|[_\-])(\d{4})(?:[_\-]|$)", pdf_path.stem)
        if m:
            inferred_year = m.group(1)
    if not inferred_year:
        for part in pdf_path.parts:
            if re.fullmatch(r"\d{4}", part):
                inferred_year = part; break

    # FIX 3: reject impossible years (typos like 2205 in filename)
    current_year = datetime.now().year
    if inferred_year and (int(inferred_year) < 1971 or int(inferred_year) > current_year + 1):
        log.warning("Suspicious year '%s' in %s — falling back to folder name", inferred_year, pdf_path.name)
        inferred_year = ""
        for part in pdf_path.parts:
            if re.fullmatch(r"\d{4}", part) and 1971 <= int(part) <= current_year + 1:
                inferred_year = part; break

    meta["year"] = int(inferred_year) if inferred_year else 0
    meta["era"]  = get_era(meta["year"]) if meta["year"] else ""

    # ── Date ──────────────────────────────────────────────────────────────────
    # FIX 2: try dot-separated date in filename first: "2002.11.2txsap.pdf" → 2002-11-02
    m_dot = re.match(r'^(\d{4})\.(\d{1,2})\.(\d{1,2})', pdf_path.stem)
    if m_dot:
        try:
            meta["date"] = datetime.strptime(
                f"{m_dot.group(1)}-{int(m_dot.group(2)):02d}-{int(m_dot.group(3)):02d}", "%Y-%m-%d"
            ).strftime("%Y-%m-%d")
        except ValueError:
            pass

    RE_DATE_NOYEAR = re.compile(
        r"(?:January|February|March|April|May|June|July|August|September|October|November|December)"
        r"[\s•·]+\d{1,2}(?:[,\-–]\s*\d{1,2})?(?!\s*,?\s*\d{4})", re.IGNORECASE,
    )
    if not meta["date"]:
        dm = RE_DATE.search(header_text)
        if dm:
            raw = dm.group(0)
            cleaned = re.sub(r"(\d+)[,\-–]\s*\d+,", r"\1,", raw)
            cleaned = re.sub(r"[•·]", " ", cleaned).strip()
            for fmt in ("%B %d, %Y", "%B %d %Y", "%m/%d/%Y"):
                try:
                    meta["date"] = datetime.strptime(cleaned, fmt).strftime("%Y-%m-%d"); break
                except ValueError: pass
            if not meta["date"] and re.match(r"\d{1,2}/\d{1,2}/\d{2}$", cleaned):
                century = inferred_year[:2] if inferred_year else "19"
                try:
                    full = re.sub(r"(\d{2})$", century + r"\1", cleaned)
                    meta["date"] = datetime.strptime(full, "%m/%d/%Y").strftime("%Y-%m-%d")
                except ValueError: pass
            if not meta["date"]:
                meta["date"] = raw
    if not meta["date"] or not re.match(r"\d{4}-\d{2}-\d{2}", meta["date"]):
        dm2 = RE_DATE_NOYEAR.search(header_text)
        if dm2 and inferred_year:
            raw2 = re.sub(r"[•·]", " ", dm2.group(0)).strip().rstrip(",")
            raw2 = re.sub(r"(\d+)[,\-–]\s*\d+\s*$", r"\1", raw2).strip()
            try:
                meta["date"] = datetime.strptime(f"{raw2}, {inferred_year}", "%B %d, %Y").strftime("%Y-%m-%d")
            except ValueError: pass
    if not meta["date"] or not re.match(r"\d{4}-\d{2}-\d{2}", meta["date"]):
        dm3 = RE_DATE_MMDD.search(header_text)
        if dm3 and inferred_year:
            try:
                meta["date"] = datetime.strptime(f"{dm3.group(0)}/{inferred_year}", "%m/%d/%Y").strftime("%Y-%m-%d")
            except ValueError: pass
    if not meta["date"] and inferred_year:
        meta["date"] = inferred_year

    # ── Performance & championship type ───────────────────────────────────────
    perf_label, panel = normalise_performance(header_text)
    # FIX 7: if header had no performance keyword, try the filename
    if not perf_label:
        stem_spaced = pdf_path.stem.replace("-", " ").replace("_", " ")
        perf_label, panel = normalise_performance(stem_spaced)
    meta["performance"] = perf_label
    meta["panel"]       = panel
    cm = RE_CHAMP.search(header_text)
    if cm:
        meta["championship_type"] = cm.group(1).title()

    # ── Competition name ───────────────────────────────────────────────────────
    RE_BOA_TITLE = re.compile(
        r"^(?:Bands of America|BOA)\s+.+?(?:Championship|Regional|Invitational|National)",
        re.IGNORECASE,
    )
    for line in lines[:15]:
        line = line.strip()
        if RE_TITLE.match(line):
            meta["raw_title"]   = line
            meta["competition"] = re.sub(r"^\d{4}\s+", "", line).strip()
            break
    if not meta["competition"]:
        for line in lines[:10]:
            line = line.strip()
            if RE_BOA_TITLE.match(line):
                meta["raw_title"]   = line
                prefix = f"{inferred_year} " if inferred_year else ""
                meta["competition"] = prefix + line
                break
    if not meta["competition"]:
        meta["competition"] = pdf_path.stem.replace("-", " ").replace("_", " ")

    # ── Venue ─────────────────────────────────────────────────────────────────
    RE_VENUE_DASH = re.compile(
        r"^(.+?)\s*[-–]\s*"
        r"(?:(?:January|February|March|April|May|June|July|August|September"
        r"|October|November|December)|\d{1,2}/\d{1,2})",
        re.IGNORECASE,
    )
    RE_DATE_ONLY = re.compile(
        r"^(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)?[,\s]*"
        r"(?:January|February|March|April|May|June|July|August|September|October|November|December)"
        r"|^\d{1,2}/\d{1,2}", re.IGNORECASE,
    )
    RE_VENUE_KEYWORDS = re.compile(
        r"stadium|dome|arena|field|park|university|college|auditorium|"
        r"center|centre|complex|gym|gymnasium|bowl|speedway|fairground",
        re.IGNORECASE,
    )
    for line in lines[:8]:
        line = line.strip()
        if not line: continue
        if RE_TITLE.match(line): continue
        if RE_BOA_TITLE.match(line): continue
        if RE_PERF.search(line) and len(line) < 40: continue
        if RE_DATE_ONLY.match(line): continue
        vm = RE_VENUE_DASH.match(line)
        if vm:
            meta["venue"] = vm.group(1).strip(); break
        if RE_VENUE_KEYWORDS.search(line):
            venue = line.split(",")[0].strip() if len(line) > 60 else line
            meta["venue"] = venue; break

    return meta


# ── Era-specific row parsers ──────────────────────────────────────────────────

def is_data_row(cells: list) -> bool:
    return sum(1 for c in cells if num_or_none(c) is not None) >= 6


def _base_record(meta: dict, source: str) -> dict:
    rec = {col: "" for col in COLUMNS}
    rec["Source File"]       = source
    rec["Date"]              = meta["date"]
    rec["Year"]              = meta["year"] or ""
    rec["Era"]               = meta["era"]
    rec["Competition"]       = meta["competition"]
    rec["Championship Type"] = meta["championship_type"]
    rec["Performance"]       = meta["performance"]
    rec["Venue Name"]        = meta.get("venue", "")
    return rec


# ─── Score field layouts ──────────────────────────────────────────────────────
#
#  CLASSIC / MID (1976-1999):
#    Ind_Mus  Ens_Mus  Avg_Mus
#    Ind_Vis  Ens_Vis  Avg_Vis
#    Mus1_GE  Mus2_GE  MusTot_GE  Vis_GE  GE_Total
#    Subtotal  Penalty  Total
#    → 14 numeric fields; NO March/Vis column distinct from Vis_GE
#
#  2000s (2000-2012): same 14-field layout, adds Rating I/II/III text token
#
#  MODERN / 2020s (2013+):
#    Some events: Ind_Mus  Ens_Mus  Avg_Mus  Ind_Vis  Ens_Vis  Avg_Vis
#                 Mus1_GE  Mus2_GE  MusTot   Vis_GE  GE_Total  Sub  Pen  Total
#    Others add a Field&Timing column instead of second GE judge
#
FIELDS_14 = [
    "Inv Mus", "Ens Mus", "Avg Mus",
    "Inv Vis", "Ens Vis", "Avg Vis",
    "Mus 1 GE", "Mus 2 GE", "Total Mus GE", "March/Vis GE", "GE Total",
    "Subtotal", "Penalty", "Total",
]
FIELDS_13 = [
    "Inv Mus", "Ens Mus", "Avg Mus",
    "Inv Vis", "Ens Vis", "Avg Vis",
    "Mus 1 GE", "Mus 2 GE", "Total Mus GE", "GE Total",
    "Subtotal", "Penalty", "Total",
]
# Modern (2013+): no March/Vis GE column.
# "GE Total" slot in these PDFs = a single standalone GE judge average (~14–18),
# NOT the summed GE from two judges as in the classic era.
# We map it into "GE Vis (Standalone)" so analysis can distinguish the two eras.
# Use this layout before FIELDS_14 for modern era to prevent the off-by-one shift.
FIELDS_13_MODERN = [
    "Inv Mus", "Ens Mus", "Avg Mus",
    "Inv Vis", "Ens Vis", "Avg Vis",
    "Mus 1 GE", "Mus 2 GE", "Total Mus GE", "GE Vis (Standalone)",
    "Subtotal", "Penalty", "Total",
]
# Some modern events have only 11 score numbers (Avg columns absent)
FIELDS_11_MODERN = [
    "Inv Mus", "Ens Mus",
    "Inv Vis", "Ens Vis",
    "Mus 1 GE", "Mus 2 GE", "Total Mus GE", "GE Vis (Standalone)",
    "Subtotal", "Penalty", "Total",
]
# Classic/mid 11-field layout (Avg columns absent, but has Vis GE sub-judge)
FIELDS_11 = [
    "Inv Mus", "Ens Mus",
    "Inv Vis", "Ens Vis",
    "Mus 1 GE", "Mus 2 GE", "Total Mus GE", "GE Total",
    "Subtotal", "Penalty", "Total",
]
# Fix 5: Classic/mid PDFs with no Avg columns → 10 numeric values per row
FIELDS_10 = [
    "Inv Mus", "Ens Mus",
    "Inv Vis", "Ens Vis",
    "Mus 1 GE", "Mus 2 GE", "GE Total",
    "Subtotal", "Penalty", "Total",
]


def _try_map_scores(fields: list, vals: list):
    """
    Slide a window of len(fields) across vals checking internal consistency.
    Returns (score_dict, start_index) or (None, -1).
    """
    n = len(fields)
    if len(vals) < n:
        return None, -1
    for start in range(len(vals) - n + 1):
        w = vals[start: start + n]
        # If layout has Avg_Mus (index 2): must be mean of Ind/Ens within tolerance
        if "Avg Mus" in fields:
            avg_idx = fields.index("Avg Mus")
            ind_idx = fields.index("Inv Mus")
            ens_idx = fields.index("Ens Mus")
            if abs(w[avg_idx] - (w[ind_idx] + w[ens_idx]) / 2) > 1.5:
                continue
        # Avg_Vis consistency
        if "Avg Vis" in fields:
            avg_idx = fields.index("Avg Vis")
            ind_idx = fields.index("Inv Vis")
            ens_idx = fields.index("Ens Vis")
            if abs(w[avg_idx] - (w[ind_idx] + w[ens_idx]) / 2) > 1.5:
                continue
        # Subtotal > GE Total  (classic/2000s era: GE Total is a large sum)
        # For modern era layouts using "GE Vis (Standalone)", Subtotal is the sum of
        # three averages (~45) which is always > the standalone GE score (~15), so
        # the check passes naturally — no separate branch needed.
        if "Subtotal" in fields and "GE Total" in fields:
            if w[fields.index("Subtotal")] <= w[fields.index("GE Total")]:
                continue
        return dict(zip(fields, w)), start
    return None, -1


def _map_scores(num_vals: list, era: str) -> tuple:
    """
    Try field layouts in priority order for the given era.
    Returns (score_dict, start_index, fields_used).

    Key fixes applied here:
      FIX 1: Modern era tries FIELDS_13_MODERN before FIELDS_14 so the single-GE-judge
              layout is matched first, preventing the off-by-one shift that left Total=0.
      FIX 5: After mapping, validate Mus2_GE <= 20; if violated, retry with FIELDS_10
              (no-average layout used by some 1980s and 2019/2025 PDFs).
    """
    if era in (ERA_CLASSIC, ERA_MID):
        priority = [FIELDS_14, FIELDS_13, FIELDS_10, FIELDS_11]
    elif era == ERA_2000s:
        priority = [FIELDS_14, FIELDS_13, FIELDS_11, FIELDS_10]
    else:
        # Modern / 2020s: try the 13-field modern layout FIRST (routes GE into
        # "GE Vis (Standalone)" and avoids the off-by-one Penalty/Total swap)
        priority = [FIELDS_13_MODERN, FIELDS_11_MODERN, FIELDS_13, FIELDS_14, FIELDS_11, FIELDS_10]

    for fields in priority:
        mapped, start = _try_map_scores(fields, num_vals)
        if mapped is None:
            continue
        # Fix 5: reject mappings where a sub-judge GE score is impossibly large
        if mapped.get("Mus 2 GE", 0) and mapped["Mus 2 GE"] > 20:
            continue
        if mapped.get("Mus 1 GE", 0) and mapped["Mus 1 GE"] > 20:
            continue
        # FIX 1: Penalty/Total swap guard.
        # Some PDFs (especially 2013+ multi-judge events) land the factored total in the
        # Penalty slot and leave Total=0. Detect by: Penalty>50, Total<5, Sub<55.
        # Swap so Total holds the score and Penalty holds the deduction (usually 0).
        pen = mapped.get("Penalty") or 0
        tot = mapped.get("Total")   or 0
        sub = mapped.get("Subtotal") or 0
        if pen > 50 and tot < 5 and sub < 55:
            mapped["Total"]   = pen
            mapped["Penalty"] = tot
        return mapped, start, fields

    # Fallback: assign positionally
    log.debug("    Score mapping fallback (positional)")
    mapped = {f: num_vals[i] if i < len(num_vals) else "" for i, f in enumerate(FIELDS_13)}
    return mapped, 0, FIELDS_13


# ── Era-specific row parsers ──────────────────────────────────────────────────

def parse_row_classic(cells: list, meta: dict, source: str):
    """
    1976–1984 format.
    Columns (after optional leading place number stripped):
      school_tokens... | Ind Mus | Ens Mus | Avg Mus | Ind Vis | Ens Vis | Avg Vis |
      Mus1 | Mus2 | MusTot | Vis GE | GE Tot | Sub | Pen | Total | Place-in-class | Class
    No Rating column. No separate state column (state embedded in school name).
    """
    rec = _base_record(meta, source)
    cells = [c.strip() for c in cells if c.strip()]

    # Strip leading rank number
    if cells and re.match(r"^\d{1,3}$", cells[0]):
        cells = cells[1:]

    num_vals  = []
    text_toks = []
    for idx, c in enumerate(cells):
        n = num_or_none(c)
        if n is not None:
            num_vals.append(n)
        else:
            text_toks.append((c, idx))

    if len(num_vals) < 6:
        return None

    mapped, score_start, used_fields = _map_scores(num_vals, ERA_CLASSIC)
    for field, val in mapped.items():
        rec[field] = val

    score_len        = len(used_fields)
    leftover_before  = [int(v) for v in num_vals[:score_start]               if v == int(v)]
    leftover_after   = [int(v) for v in num_vals[score_start + score_len:]   if v == int(v)]

    place_overall = ""
    place_class   = ""

    if leftover_before:
        if meta["panel"]:
            rec["Place in Panel"] = str(leftover_before[0])
        else:
            place_overall = str(leftover_before[0])

    # Classic format: "Place in Class" then "Class" appear AFTER scores as text tokens
    # Numeric leftovers after scores are place-in-class numbers
    if len(leftover_after) == 1:
        place_class = str(leftover_after[0])
    elif len(leftover_after) >= 2:
        place_class   = str(leftover_after[0])
        place_overall = str(leftover_after[1])

    first_num_cell = next(
        (idx for idx, c in enumerate(cells) if num_or_none(c) is not None), len(cells)
    )
    pre_text  = [c for c, i in text_toks if i < first_num_cell]
    post_text = [c for c, i in text_toks if i >= first_num_cell]

    school_toks = list(pre_text)
    class_label = ""
    for tok in post_text:
        if RE_CLASS.match(tok):
            class_label = tok.upper()
        elif not RE_NUM.match(tok) and not RE_ROMAN.match(tok) and not class_label:
            school_toks.append(tok)

    school, state, ext_penalty = parse_school_and_state(school_toks)
    rec["School"]         = school
    rec["School State"]   = state
    rec["Place in Class"] = place_class
    rec["Place Overall"]  = place_overall
    rec["Class"]          = class_label
    rec["Rating"]         = ""   # No rating in classic era
    if ext_penalty is not None and not rec["Penalty"]:
        rec["Penalty"] = ext_penalty

    if not rec["Place Overall"] and rec["Place in Panel"]:
        rec["Place Overall"] = rec["Place in Panel"]
    return rec


def parse_row_mid(cells: list, meta: dict, source: str):
    """
    1985–1999 format. Mostly identical to classic but may carry a Rating
    token (I / II / III or 1/2/3) after the Class label.
    """
    rec = parse_row_classic(cells, meta, source)
    if rec is None:
        return None
    # Re-examine post-score tokens for Rating
    cells_clean = [c.strip() for c in cells if c.strip()]
    if cells_clean and re.match(r"^\d{1,3}$", cells_clean[0]):
        cells_clean = cells_clean[1:]
    first_num_cell = next(
        (i for i, c in enumerate(cells_clean) if num_or_none(c) is not None), len(cells_clean)
    )
    post_text = [cells_clean[i] for i in range(first_num_cell, len(cells_clean))
                 if num_or_none(cells_clean[i]) is None]
    found_class = bool(rec.get("Class"))
    for tok in post_text:
        if RE_ROMAN.match(tok):
            rec["Rating"] = tok; break
        if RE_RATING_DIGIT.match(tok) and found_class:
            rec["Rating"] = tok; break
    return rec


def parse_row_2000s(cells: list, meta: dict, source: str):
    """
    2000–2012 format.
    Adds a separate State column (two-letter) after school name in many PDFs.
    Rating I/II/III appears at the end.
    Overall place appears as an extra integer at the end.
    Layout: [rank] school [state] scores... [Place-in-class] [Place-overall] [Rating]
    """
    rec = _base_record(meta, source)
    cells = [c.strip() for c in cells if c.strip()]

    if cells and re.match(r"^\d{1,3}$", cells[0]):
        cells = cells[1:]

    num_vals  = []
    text_toks = []
    for idx, c in enumerate(cells):
        n = num_or_none(c)
        if n is not None:
            num_vals.append(n)
        else:
            text_toks.append((c, idx))

    if len(num_vals) < 6:
        return None

    mapped, score_start, used_fields = _map_scores(num_vals, ERA_2000s)
    for field, val in mapped.items():
        rec[field] = val

    score_len       = len(used_fields)
    leftover_before = [int(v) for v in num_vals[:score_start]              if v == int(v)]
    leftover_after  = [int(v) for v in num_vals[score_start + score_len:]  if v == int(v)]

    place_overall = ""
    place_class   = ""

    if leftover_before:
        place_overall = str(leftover_before[0])

    if len(leftover_after) == 1:
        if not place_overall:
            place_overall = str(leftover_after[0])
        else:
            place_class = str(leftover_after[0])
    elif len(leftover_after) >= 2:
        place_class   = str(leftover_after[0])
        place_overall = str(leftover_after[1])

    first_num_cell = next(
        (idx for idx, c in enumerate(cells) if num_or_none(c) is not None), len(cells)
    )
    pre_text  = [c for c, i in text_toks if i < first_num_cell]
    post_text = [c for c, i in text_toks if i >= first_num_cell]

    school_toks = list(pre_text)
    class_label = ""
    rating      = ""
    found_class = False
    for tok in post_text:
        if RE_ROMAN.match(tok):
            rating = tok
        elif RE_CLASS.match(tok):
            class_label = tok.upper(); found_class = True
        elif RE_RATING_DIGIT.match(tok) and found_class:
            rating = tok
        elif not RE_NUM.match(tok) and not RE_ROMAN.match(tok) and not found_class:
            school_toks.append(tok)

    school, state, ext_penalty = parse_school_and_state(school_toks)
    rec["School"]         = school
    rec["School State"]   = state
    rec["Place in Class"] = place_class
    rec["Place Overall"]  = place_overall
    rec["Class"]          = class_label
    rec["Rating"]         = rating
    if ext_penalty is not None and not rec["Penalty"]:
        rec["Penalty"] = ext_penalty
    return rec


def parse_row_modern(cells: list, meta: dict, source: str):
    """
    2013–2019 format (CompetitionSuite PDFs).
    School format: "Name H.S. - ST"
    Columns after scores: Subtotal | Penalty | Total | Overall
    No explicit Class column on many events (it's encoded in the section header).
    Rating appears at end as I/II/III on some events (pre-2017),
    then is dropped and replaced by Overall rank.
    """
    rec = _base_record(meta, source)
    cells = [c.strip() for c in cells if c.strip()]

    if cells and re.match(r"^\d{1,3}$", cells[0]):
        cells = cells[1:]

    num_vals  = []
    text_toks = []
    for idx, c in enumerate(cells):
        n = num_or_none(c)
        if n is not None:
            num_vals.append(n)
        else:
            text_toks.append((c, idx))

    if len(num_vals) < 6:
        return None

    mapped, score_start, used_fields = _map_scores(num_vals, ERA_MODERN)
    for field, val in mapped.items():
        rec[field] = val

    score_len       = len(used_fields)
    leftover_before = [int(v) for v in num_vals[:score_start]              if v == int(v)]
    leftover_after  = [int(v) for v in num_vals[score_start + score_len:]  if v == int(v)]

    place_overall = ""
    place_class   = ""

    if leftover_before:
        place_overall = str(leftover_before[0])

    if len(leftover_after) == 1:
        if not place_overall:
            place_overall = str(leftover_after[0])
    elif len(leftover_after) >= 2:
        place_class   = str(leftover_after[0])
        place_overall = str(leftover_after[1])

    first_num_cell = next(
        (idx for idx, c in enumerate(cells) if num_or_none(c) is not None), len(cells)
    )
    pre_text  = [c for c, i in text_toks if i < first_num_cell]
    post_text = [c for c, i in text_toks if i >= first_num_cell]

    school_toks = list(pre_text)
    class_label = ""
    rating      = ""
    found_class = False
    for tok in post_text:
        if RE_ROMAN.match(tok):
            rating = tok
        elif RE_CLASS.match(tok):
            class_label = tok.upper(); found_class = True
        elif RE_RATING_DIGIT.match(tok) and found_class:
            rating = tok
        elif not RE_NUM.match(tok) and not RE_ROMAN.match(tok) and not found_class:
            school_toks.append(tok)

    school, state, ext_penalty = parse_school_and_state(school_toks)
    rec["School"]         = school
    rec["School State"]   = state
    rec["Place in Class"] = place_class
    rec["Place Overall"]  = place_overall
    rec["Class"]          = class_label
    rec["Rating"]         = rating
    if ext_penalty is not None and not rec["Penalty"]:
        rec["Penalty"] = ext_penalty
    return rec


# 2020s format is functionally identical to Modern; alias it
parse_row_2020s = parse_row_modern


# ── Dispatcher: pick era parser ───────────────────────────────────────────────

ERA_PARSER = {
    ERA_CLASSIC: parse_row_classic,
    ERA_MID:     parse_row_mid,
    ERA_2000s:   parse_row_2000s,
    ERA_MODERN:  parse_row_modern,
    ERA_2020s:   parse_row_2020s,
}


def parse_pdf(pdf_path: Path) -> list:
    text = extract_text(pdf_path)
    if not text:
        log.warning("No text extracted: %s", pdf_path)
        return []

    lines = [l for l in text.splitlines() if l.strip()]
    meta  = parse_header(lines, pdf_path)
    era   = meta["era"] or ERA_2020s   # default to modern if unknown

    row_parser = ERA_PARSER.get(era, parse_row_modern)

    log.info("  Parsing: %s  [%s | era=%s | %s | panel=%s]",
             pdf_path.name, meta["date"], era, meta["performance"], meta["panel"] or "—")

    table_rows = extract_table_rows(pdf_path)
    records    = []

    for row in table_rows:
        if not is_data_row(row):
            continue
        rec = row_parser(row, meta, str(pdf_path))
        if rec and rec.get("School"):
            records.append(rec)

    # Deduplicate
    seen, unique = set(), []
    for r in records:
        key = (r["Date"], r["Competition"], r["Performance"], r["School"])
        if key not in seen:
            seen.add(key); unique.append(r)

    # FIX 9: compute Avg Mus / Avg Vis when blank (PDFs that omit the average column)
    for r in unique:
        if r["Avg Mus"] in (None, "") and r["Inv Mus"] not in (None, "") and r["Ens Mus"] not in (None, ""):
            try:
                r["Avg Mus"] = round((float(r["Inv Mus"]) + float(r["Ens Mus"])) / 2, 2)
            except (TypeError, ValueError):
                pass
        if r["Avg Vis"] in (None, "") and r["Inv Vis"] not in (None, "") and r["Ens Vis"] not in (None, ""):
            try:
                r["Avg Vis"] = round((float(r["Inv Vis"]) + float(r["Ens Vis"])) / 2, 2)
            except (TypeError, ValueError):
                pass

    # FIX 8: For single-class events, if Place in Class is set but Place Overall is blank,
    # copy Place in Class → Place Overall (classic/mid era events with no overall rank)
    # Conversely, if Place Overall is set and Place in Class is blank, copy the other way.
    for r in unique:
        if r["Place in Class"] and not r["Place Overall"]:
            # Only copy if all bands appear to be in one class (no class label = single-class event)
            # We can't check other rows here so we do it conservatively:
            # if there's no Class label at all, treat place-in-class as overall
            if not r["Class"]:
                r["Place Overall"] = r["Place in Class"]
        elif r["Place Overall"] and not r["Place in Class"]:
            if not r["Class"]:
                r["Place in Class"] = r["Place Overall"]

    log.info("    → %d records  (era: %s)", len(unique), era)
    return unique


# ── Excel writer ──────────────────────────────────────────────────────────────

# Era colour coding for the Year/Era columns
ERA_COLORS = {
    ERA_CLASSIC: "C6EFCE",   # green
    ERA_MID:     "FFEB9C",   # yellow
    ERA_2000s:   "BDD7EE",   # blue
    ERA_MODERN:  "E2EFDA",   # light green
    ERA_2020s:   "FCE4D6",   # orange
}


def write_excel(all_records: list, output_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recaps"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    dat_font  = Font(name="Arial", size=9)
    c_align   = Alignment(horizontal="center", vertical="center")
    l_align   = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill  = PatternFill("solid", fgColor="EEF3F8")

    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border

    numeric_cols = {
        "Inv Mus","Ens Mus","Avg Mus",
        "Inv Vis","Ens Vis","Avg Vis",
        "Mus 1 GE","Mus 2 GE","Total Mus GE","March/Vis GE","GE Total",
        "GE Vis (Standalone)",
        "Subtotal","Penalty","Total",
    }
    integer_cols = {"Place in Class","Place Overall","Place in Panel","Year"}
    left_cols    = {"School","Competition","Championship Type","Performance",
                    "Date","Source File","Venue Name","Street Address","City","Era"}

    for ri, rec in enumerate(all_records, 2):
        era   = rec.get("Era", "")
        era_color = ERA_COLORS.get(era, "FFFFFF")
        era_fill  = PatternFill("solid", fgColor=era_color)
        row_fill  = alt_fill if ri % 2 == 0 else None

        for ci, col in enumerate(COLUMNS, 1):
            val = rec.get(col, "")
            if col in numeric_cols and val != "":
                try:    val = round(float(val), 2)
                except: pass
            elif col in integer_cols and val != "":
                try:    val = int(val)
                except: pass

            val  = sanitize(val)
            cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
            cell.font      = dat_font
            cell.border    = border
            cell.alignment = l_align if col in left_cols else c_align
            if col in ("Year", "Era"):
                cell.fill = era_fill
            elif row_fill:
                cell.fill = row_fill
            if col in numeric_cols:
                cell.number_format = "0.00"

    widths = {
        "Date":18,"Year":7,"Era":20,
        "Competition":42,"Championship Type":18,"Performance":14,
        "Venue Name":30,"Street Address":35,"City":20,"State":8,"Zip":10,
        "School":35,"School State":8,
        "Inv Mus":8,"Ens Mus":8,"Avg Mus":8,
        "Inv Vis":8,"Ens Vis":8,"Avg Vis":8,
        "Mus 1 GE":9,"Mus 2 GE":9,"Total Mus GE":11,"March/Vis GE":11,"GE Total":9,
        "GE Vis (Standalone)":16,
        "Subtotal":9,"Penalty":8,"Total":9,
        "Place in Class":13,"Class":8,"Place Overall":13,"Place in Panel":13,
        "Rating":8,"Source File":50,
    }
    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 10)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Per-era summary sheets ─────────────────────────────────────────────────
    era_map: dict = {}
    for r in all_records:
        era_map.setdefault(r.get("Era","Unknown"), []).append(r)

    for era_name, recs in sorted(era_map.items()):
        ws_era = wb.create_sheet(era_name[:31])   # sheet name max 31 chars
        for ci, col in enumerate(COLUMNS, 1):
            cell = ws_era.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = hdr_align; cell.border = border
        for ri, rec in enumerate(recs, 2):
            for ci, col in enumerate(COLUMNS, 1):
                val = rec.get(col, "")
                if col in numeric_cols and val != "":
                    try:    val = round(float(val), 2)
                    except: pass
                val  = sanitize(val)
                cell = ws_era.cell(row=ri, column=ci, value=val if val != "" else None)
                cell.font = dat_font; cell.border = border
                cell.alignment = l_align if col in left_cols else c_align
                if col in numeric_cols:
                    cell.number_format = "0.00"
        for ci, col in enumerate(COLUMNS, 1):
            ws_era.column_dimensions[get_column_letter(ci)].width = widths.get(col, 10)
        ws_era.freeze_panes = "A2"
        ws_era.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Music For All Recap Database"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    years = [r["Year"] for r in all_records if r.get("Year")]
    summary_rows = [
        ("Total band performances:", len(all_records)),
        ("Unique competitions:",     len({r["Competition"] for r in all_records})),
        ("Year range:",              f"{min(years)} – {max(years)}" if years else "N/A"),
        ("Generated:",               datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Era", "Records"),
    ] + [(k, len(v)) for k, v in sorted(era_map.items())]

    for i, (label, val) in enumerate(summary_rows, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
        ws2.cell(row=i, column=2, value=val  ).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    log.info("Saved: %s  (%d rows)", output_path, len(all_records))


# ── File discovery ─────────────────────────────────────────────────────────────

NON_RECAP_PATTERNS = re.compile(
    r"map|schedule|guide|program|parking|expo|handbook|bracket|"
    r"souvenir|flyer|promo|info|instruction|registration|order|"
    r"ticket|waiver|contract|invoice|logo|cover|welcome|agenda",
    re.IGNORECASE,
)


def is_recap_pdf(pdf_path: Path) -> bool:
    return not NON_RECAP_PATTERNS.search(pdf_path.stem)


def find_pdfs(input_dir: Path) -> list:
    all_pdfs    = sorted(input_dir.rglob("*.pdf"))
    recap_pdfs  = [p for p in all_pdfs if is_recap_pdf(p)]
    skipped     = len(all_pdfs) - len(recap_pdfs)
    if skipped:
        log.info("Skipping %d non-recap PDFs", skipped)

    seen_names: dict = {}
    for p in recap_pdfs:
        name = p.name.lower()
        if name not in seen_names:
            seen_names[name] = p
        elif len(p.parts) > len(seen_names[name].parts):
            seen_names[name] = p

    deduped = sorted(seen_names.values())
    dupes   = len(recap_pdfs) - len(deduped)
    if dupes:
        log.info("Removed %d duplicate PDF paths", dupes)
    return deduped


# ── Main ──────────────────────────────────────────────────────────────────────

DEFAULT_INPUT  = Path("musicforall_recaps")
DEFAULT_OUTPUT = Path("recaps.xlsx")


def main():
    parser = argparse.ArgumentParser(
        description="Parse Music For All recap PDFs → Excel (era-segmented)"
    )
    parser.add_argument("--input",  default=str(DEFAULT_INPUT),
                        help="Root folder containing year sub-folders of PDFs")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT),
                        help="Output Excel file path")
    parser.add_argument("--year",   type=int, default=None,
                        help="Process only this specific year (e.g. --year 1995)")
    parser.add_argument("--era",    default=None,
                        help="Process only a specific era name substring "
                             "(e.g. --era Classic)")
    parser.add_argument("--no-geocode", action="store_true",
                        help="Skip venue geocoding")
    args = parser.parse_args()

    input_dir   = Path(args.input)
    output_path = Path(args.output)

    if not input_dir.exists():
        log.error("Input folder not found: %s", input_dir)
        sys.exit(1)

    pdfs = find_pdfs(input_dir)

    # Optional year / era filter
    if args.year:
        year_str = str(args.year)
        pdfs = [p for p in pdfs
                if any(part == year_str for part in p.parts)
                or p.stem.startswith(year_str)]
        log.info("Year filter %d → %d PDFs", args.year, len(pdfs))

    log.info("Found %d PDFs in %s", len(pdfs), input_dir)

    all_records = []
    total_pdfs  = len(pdfs)
    for i, pdf_path in enumerate(pdfs, 1):
        if i % 50 == 0 or i == 1 or i == total_pdfs:
            print(f"  Progress: {i}/{total_pdfs} PDFs  ({len(all_records)} records so far)...")
        try:
            recs = parse_pdf(pdf_path)
            # Optional era filter
            if args.era:
                recs = [r for r in recs if args.era.lower() in r.get("Era","").lower()]
            all_records.extend(recs)
        except Exception as exc:
            log.error("Failed: %s — %s", pdf_path, exc)

    log.info("Total records: %d", len(all_records))
    if not all_records:
        log.warning("No records found. Check parse_log.txt")
        sys.exit(0)

    # ── Geocode ───────────────────────────────────────────────────────────────
    if not args.no_geocode:
        geocache = load_geocache()
        unique_venues = {
            (r["Venue Name"], r["Competition"])
            for r in all_records if r.get("Venue Name")
        }
        log.info("Geocoding %d unique venues...", len(unique_venues))
        venue_results: dict = {}
        for venue_name, competition in sorted(unique_venues):
            geo = geocode_venue(venue_name, competition, geocache)
            venue_results[(venue_name, competition)] = geo
            import time; time.sleep(0.2)
        save_geocache(geocache)
        for r in all_records:
            key = (r.get("Venue Name",""), r.get("Competition",""))
            geo = venue_results.get(key, {})
            r["Street Address"] = geo.get("street","")
            r["City"]           = geo.get("city",  "")
            r["State"]          = geo.get("state", "")
            r["Zip"]            = geo.get("zip",   "")
    else:
        log.info("Geocoding skipped (--no-geocode)")

    # ── Sort ──────────────────────────────────────────────────────────────────
    def sort_key(r):
        date  = r.get("Date","") or ""
        comp  = r.get("Competition","") or ""
        perf  = r.get("Performance","") or ""
        panel = r.get("Place in Panel","") or "0"
        try:    place = int(r.get("Place Overall") or r.get("Place in Panel") or 9999)
        except: place = 9999
        return (date, comp, perf, panel, place)

    all_records.sort(key=sort_key)
    write_excel(all_records, output_path)

    # Print era breakdown to console
    from collections import Counter
    era_counts = Counter(r.get("Era","?") for r in all_records)
    print(f"\n✓ Done! {len(all_records)} total rows → {output_path.resolve()}")
    print("\n  Records by era:")
    for era_name, count in sorted(era_counts.items()):
        print(f"    {era_name:<30} {count:>5}")


if __name__ == "__main__":
    main()
